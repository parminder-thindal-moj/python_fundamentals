from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def read_workbook(file_path):

    wb = load_workbook(filename=file_path)

    sheet_names = wb.sheetnames

    return wb, sheet_names


def print_sheet_names(sheet_names):
    for sheet in sheet_names:
        print(sheet)


# Guard clause - inverts the previous code.
def create_intro_sheet(wb, sheet_names):
    if "Intro" not in sheet_names:
        intro_sheet = wb.create_sheet("Intro", 0)
    else:
        intro_sheet = wb["Intro"]

    return intro_sheet


def customise_intro_sheet(wb, intro_sheet, sheet_names):

    intro_sheet["A1"] = "JAS Metadata Overview"
    intro_sheet["A1"].font = Font(name="Arial", size=32, bold=True)

    # add headers:
    intro_sheet["A5"] = "Sheets"
    intro_sheet["A5"].font = Font(name="Arial", size=16, bold=True)

    intro_sheet["B5"] = "Info"
    intro_sheet["B5"].font = Font(name="Arial", size=16, bold=True)

    for idx, sheet_name in enumerate(sheet_names, start=7):
        if sheet_name != "Intro":

            # Create the hyperlink in column A that links to the sheet
            intro_sheet[f"A{idx}"] = f"{sheet_name}"

            intro_sheet[f"A{idx}"].hyperlink = (
                f"#{sheet_name}!A1"  # Internal link to the sheet
            )

            intro_sheet[f"A{idx}"].style = "Hyperlink"  # Apply hyperlink style


def resize_columns(wb, intro_sheet):

    # Automatically resize column A
    max_length = 0
    for row in intro_sheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error measuring cell length: {e}")

    intro_sheet.column_dimensions["A"].width = max_length + 5  # Adding padding

    return intro_sheet


# Save the updated workbook
def output_wb(wb, file_path):
    wb.save(file_path)


if __name__ == "__main__":

    file_path = "my_file_path.csv"

    wb, sheet_names = read_workbook(file_path=file_path)

    intro_sheet = create_intro_sheet(wb=wb, sheet_names=sheet_names)

    customise_intro_sheet(wb=wb, intro_sheet=intro_sheet, sheet_names=sheet_names)

    print_sheet_names(sheet_names=sheet_names)

    output_wb(wb=wb, file_path=file_path)

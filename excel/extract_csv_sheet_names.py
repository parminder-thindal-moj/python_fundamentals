from openpyxl import load_workbook

wb = load_workbook(
    "/Users/parminder.thindal/Library/CloudStorage/OneDrive-MinistryofJustice/Cross-CJS data asset/Official Alpha-CJS documentation/Metadata/Alpha-CJS Phase 2 metadata.xlsx"
)

sheet_names = wb.sheetnames

# if "Intro" in sheet_names:
#     intro_sheet = wb['Intro']
# else:
#     intro_sheet = wb.create_sheet('Intro')

# Guard clause - inverts the previous code.
if "Intro" not in sheet_names:
    intro_sheet = wb.create_sheet("Intro", 0)
else:
    intro_sheet = wb["Intro"]


if __name__ == "__main__":
    for sheet in sheet_names:
        print(sheet)
